from openpyxl import load_workbook
from pprint import pprint
from requests import post

# 用例Excel名称
DATA_PATH = "test_case_api.xlsx"
# 用例sheet表名称,当前需要执行的模块
# SHEET_NAME = "register"
SHEET_NAME = "login"


# 获取接口用例
def read_data(data_path, sheet_name):
    """
    读取接口用例，并返回用例列表
    data_path：Excel名字，需要放入当前py文件目录
    sheet_name：需要执行测试的用例所在sheet表名
    """
    # 打开用例Excel
    wb = load_workbook(filename=data_path)
    # 获取用例所在表
    sheet = wb[sheet_name]
    # 获取最大行号
    _max_row = sheet.max_row
    # 定义返回的列表
    cases_list = []
    # 循环获取测试用例，并存以字典形式入cases_list
    for _row in range(2, _max_row + 1):
        # 字典的第二种定义方法dict(key=value)==>{"key": "value"}
        case_id = sheet.cell(row=_row, column=1).value
        print(case_id)
        dic = dict(
                case_id = sheet.cell(row=_row, column=1).value,
                case_url = sheet.cell(row=_row, column=5).value,
                # eval执行字符串里的可执行代码，这里是为了将字符串变成字典格式
                case_json = eval(sheet.cell(row=_row, column=6).value),
                case_headers = eval(sheet.cell(row=_row, column=8).value),
                case_expected = eval(sheet.cell(row=_row, column=7).value)
        )
        # 将每一行的用例字典存入列表
        cases_list.append(dic)
    # 返回用例列表
    return cases_list


# 访问接口函数
def QCD_IT_Fiddler(qcd_url, content, qcd_headers):
    """
    前程贷接口测试入口,返回json响应体
    qcd_url：接口地址
    content：json格式的请求体
    qcd_headers：字典格式的请求头部信息
    """
    # 访问具体接口，接收返回信息
    response = post(url=qcd_url, json=content, headers=qcd_headers)
    # 函数返回json格式的响应体
    return response.json()


# 批量执行并返回执行结果
def perform_IT(case_list):
    """
    执行接口测试，返回字典格式行号和断言 key--行号 value：断言
    case_list：需要执行的用例，列表存字典的格式
    """
    # 定义返回的字典
    perform_result = {}
    # 定义起始行号，一般表格会定义表头，所以都是从第二行开始
    active_row = 2
    # 循环执行每行测试用例，并断言，将行号，断言存入字典
    for case in case_list:
        # 接口地址
        url = case.get("case_url")
        # 请求体
        json_data = case.get("case_json")
        # 请求头
        header_data = case.get("case_headers")
        # 期望
        expected_msg = case.get("case_expected").get("msg")
        # 执行接口测试，并接收返回体
        result_register = QCD_IT_Fiddler(qcd_url=url, content=json_data, qcd_headers=header_data)

        # 断言
        if expected_msg == result_register.get("msg"):
            # 成功
            perform_result.update({active_row: "Passed"})
        else:
            # 失败
            perform_result.update({active_row: "Failed"})
        # 行号+1
        active_row += 1
    # 返回执行结果
    return perform_result


# 将结果写入对应表格
def write_result(data_path, sheet_name, result_dic):
    """
    将执行结果写入Excel的result列
    data_path：执行的用例Excel名称
    sheet_name：执行的用例表格名称
    result_dic：执行结果，字典格式的{行号：断言,...}
    """
    wb = load_workbook(filename=data_path)
    sheet = wb[sheet_name]
    # 以key作为行号，行号存入的断言作为执行结果写入表格的result列
    for _row in result_dic.keys():
        sheet.cell(row=_row, column=9).value = result_dic.get(_row)
    # 保存数据
    wb.save(data_path)


# 获取接口测试用例,用cases接收
cases = read_data(DATA_PATH, SHEET_NAME)
# 执行接口测试
result = perform_IT(cases)
# 将执行结果写入Excel
write_result(DATA_PATH, SHEET_NAME, result)